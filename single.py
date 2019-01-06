# coding: utf-8
from sentiment_analysis_test import setiment_score
import openpyxl
import threading 
import sys 

book = openpyxl.load_workbook('./weibo_data_v0.xlsx')
sheet = book.active

comments = []
for row in sheet.iter_rows(min_row=1, min_col=6, max_row=sheet.max_row,  max_col=6):
    for cell in row:
        if not cell.value:
            continue 
        comments.append(cell.value)

print 'load done'

print sys.argv[1],sys.argv[2],sys.argv[3]

start = int(sys.argv[1])
end  = int(sys.argv[2])

print start, end

fname = sys.argv[3]

f = open(fname, "w")
for i in range(start, end):
    output = str(setiment_score(comments[i])) + '\n'
    f.write(output)
f.close()


print sys.argv[3], ' done'
