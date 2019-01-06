# coding: utf-8
from sentiment_analysis_test import setiment_score
import openpyxl
import threading 

book = openpyxl.load_workbook('./weibo_data_v0.xlsx')
sheet = book.active

comments = []
for row in sheet.iter_rows(min_row=1, min_col=6, max_row=sheet.max_row,  max_col=6):
    for cell in row:
        if not cell.value:
            continue 
        comments.append(cell.value)

print comments[0]
def analysis_segment(start, end, i):
    f = open('part-'+str(i)+'.txt', "w")
    for i in range(start, end):
        output = str(setiment_score(comments[i]))
        f.write(output)
    f.close()
threads = []
NUM_threads = 8
BEGIN = 1
END = len(comments)-1
SEG = int(float(END)/NUM_threads)+1
for i in range(0, NUM_threads):
    start = i* SEG
    end = min(start+SEG, END)
    print start, end
    t = threading.Thread(target = analysis_segment, args = (start, end, i))
    threads.append(t)
    t.start()

for t in threads:
    t.join()

