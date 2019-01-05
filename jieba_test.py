
# coding: utf-8
import jieba
import jieba.posseg as pseg
import openpyxl
from openpyxl import Workbook


def cut_word():
    book = openpyxl.load_workbook('./weibo_data_v0.xlsx')

    sheet = book.active

    comments = []

    cut_res = Workbook()
    res_sheet = cut_res.active

    for row in sheet.iter_rows(min_row=1, min_col=6, max_row=sheet.max_row,  max_col=6):
        for cell in row:
            comment = cell.value
            if not comment:
                continue
            words = ' '.join(jieba.cut(comment))
            words = str(words.encode('utf-8'))
            res_sheet.append([words])

    cut_res.save('word_cut_results.xlsx')

def load_stop_word():
    stop = [line.strip().decode('utf-8') for line in open('stopwords.txt').readlines() ]
    return stop
def dump_stop(stop):
    for sw in stop:
        sw = u''.join(sw)
        print sw

def dump(word):
    print u''.join(word)

def remove_stop_and_get_flag(sentence, stop, flag):
    final = u''
    final_with_flag = u''
    segs = pseg.cut(sentence)
    for seg, flag in segs:
        if seg not in stop:
            if flag != 'm' and flag != 'x':
                final += u' '+seg
                final_with_flag += u' '+seg+'/'+flag
    #print "final is ", final
    #print "final with flag is: ", final_with_flag
    #return final_with_flag
    if flag:
        return final_with_flag
    else:
        return final

def remove_stop(output, with_flag):
    stop_words = load_stop_word()
    book = openpyxl.load_workbook('word_cut_results.xlsx')
    sheet = book.active
    count = 0 
    res = Workbook()
    res_sheet = res.active
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row,  max_col=1):
        for cell in row:
            if not cell.value:
                continue
            ret =  remove_stop_and_get_flag(cell.value, stop_words, with_flag)
            res_sheet.append([ret])

    res.save(output)
def main():
    #不带词性
    #remove_stop('remove_stop_words_with_flag.xlsx', 1)
    #带词性
    remove_stop('remove_stop_words_without_flag.xlsx', 0)
if __name__ == "__main__":
    main()










	

 

 	
